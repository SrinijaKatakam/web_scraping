from logging import exception

from bs4 import BeautifulSoup as BS
import pandas as pd
import  re
from openpyxl import load_workbook
import time
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC, expected_conditions

DRIVER_PATH = r'C:\Users\Srinija Katakam\Downloads\chromedriver_win32/chromedriver'
driver = webdriver.Chrome(DRIVER_PATH)


def extract(url):
    driver.get(url)

    try:
        pathelement = driver.find_element_by_xpath("//*[contains(text(), 'Engagement')]").find_element_by_xpath("./..") \
        .find_element_by_xpath("./..").get_attribute('id')
        print("id",pathelement)
        e = driver.find_element_by_xpath('//*[@id="' + pathelement + '"]')
        e.click()
        time.sleep(10)
    except:
        print('No engagement module in web page')
    #driver.implicitly_wait(10000)

    soup = BS(driver.page_source, "html.parser")


    return soup


def transform(soup):

    try:
        table = soup.find("div", {"class": "page-layout-body layout-column layout-align-space-between"})
        for row in table.findAll('div',
                                 attrs={'class': 'info'}):
            names = row.find('span', attrs={'class': 'wrappable-label-with-info ng-star-inserted'})
            titles = names.find('span', attrs={'_ngcontent-sc419': ''})
            values = row.find('span',
                              attrs={'class': re.compile('component--field-formatter field-type-*')})
            quote[titles.text] = values.text
    except:
        print("No Highlights in here ")

    try:
        table = soup.find('a', {'title': 'BuiltWith'})
        parent = table.parent
        values = parent.find('span', attrs={'class': re.compile('component--field-formatter field-type-*')})
        quote['BuiltWith'] = values.text
    except:
        quote['BuiltWith'] = "None"

    try:
        patent = soup.find('a', attrs={'href': re.compile(r'.patent*')})
        quote['Patent'] = patent.text
    except:
        quote['Patent'] = "None"
    try:
        trademark = soup.find('a', attrs={'href': re.compile(r'.trademark*')})
        quote['TradeMark'] = trademark.text
    except:
        quote['TradeMark'] = "None"
    try:
        ITSpend = soup.find('a', attrs={'href': re.compile(r'.aberdeen_site_it_spend*')})
        quote['ITSpend'] = ITSpend.text
    except:
        quote['ITSpend'] = "None"

    return quote

if __name__ == '__main__':
    cmpOutput = []
    start = time.time()
    df = pd.read_excel('Book1.xlsx', sheet_name='Sheet1')  # can also index sheet by name or fetch all sheets
    wb = load_workbook(filename='Book1.xlsx',
                       read_only=True)

    ws = wb['Sheet1']
    data_rows = []
    for row in ws['A2':'A101']:
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)
    df = pd.DataFrame(data_rows)
    print(df)
    print(df[0])
    mylist = df[0].tolist()
    for i in mylist:
        quote = {}
        print(i)
        quote['URL']=i
        cmp = extract(i)
        #driver.implicitly_wait(10000)
        cmpInfoList = transform(cmp)
        try:
            time.sleep(10)
            pathelement = driver.find_element_by_xpath("//*[contains(text(), 'Engagement')]").find_element_by_xpath(
                "./..").find_element_by_xpath("./..").get_attribute('id')
            print("pathelement",pathelement)
            e = driver.find_element_by_xpath('//*[@id="' + pathelement + '"]')
            e.click()
            time.sleep(10)
            labelnumber = pathelement.split("label-", 1)[1]
            print("labelnumber", labelnumber)
            element1 = driver.find_element_by_xpath("//*[@id='mat-tab-content-" + labelnumber + "']")
            print("Element1 Class:", element1.get_attribute('class'))
            print("Element1 Text", element1.text)
            ElementsString = element1.text
            order = ElementsString.replace("\n", "!").strip()
            print(order)
            my_list = order.split("!")
            print(my_list)
            stripped = [s.strip() for s in my_list]
            print("stripped:", stripped)
            print(len(stripped))
            list_length = len(stripped)
            for i in range(0, list_length - 1, 2):
                quote[stripped[i]] = stripped[i + 1]

            print("quote",quote)
        except:
            print('No values in engagement module')

        cmpOutput.append(dict(list(cmpInfoList.items())))
    df = pd.DataFrame(cmpOutput)
    print(df)
    elapsed_Time = time.time() - start
    print(elapsed_Time)
    df.to_csv('crunchbasetask11.csv')
